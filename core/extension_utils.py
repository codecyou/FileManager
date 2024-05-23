# 扩展工具模块，存放一些引用第三方库的函数代码
import time
import os
from datetime import datetime
from win32file import CreateFile, SetFileTime, GetFileTime, CloseHandle, CreateDirectory
from win32file import GENERIC_READ, GENERIC_WRITE, OPEN_EXISTING, FILE_FLAG_BACKUP_SEMANTICS
import win32api
import win32timezone  # 若源码不导入，则pyinstaller打包时不会打包该模块，会导致运行修改时间戳函数时缺少该模块而报错
import openpyxl


def get_timestr(timestamp):
    """转换时间戳为标准时间字符串"""
    return datetime.strptime(time.strftime(r"%Y-%m-%d %H:%M:%S", time.localtime(timestamp)), r"%Y-%m-%d %H:%M:%S")


def modifyFileTimeByTimestamp(path, createTimestamp, modifyTimestamp, accessTimestamp):
    """
    用来修改任意文件的相关时间属性，通过时间戳
    :param path: 文件路径名
    :param createTimeamp: 创建时间戳
    :param modifyTimestamp: 修改时间戳
    :param accessTimestamp: 访问时间戳
    :param
    """
    try:
        # 在CreateFile中仅需传入额外的标志位FILE_FLAG_BACKUP_SEMANTICS即可打开目录。在经过测试也的确能够使用CreateFile打开目录并用SetFileTime对其设置时间。
        # 官方文档 https://learn.microsoft.com/zh-cn/windows/win32/api/fileapi/nf-fileapi-createfilea#Directories
        if(os.path.isfile(path)):
            fh = CreateFile(path, GENERIC_READ | GENERIC_WRITE, 0, None, OPEN_EXISTING, 0, 0)
        elif (os.path.isdir(path)):
            fh = CreateFile(path, GENERIC_READ | GENERIC_WRITE, 0, None, OPEN_EXISTING, FILE_FLAG_BACKUP_SEMANTICS , 0)

        createTimes = get_timestr(createTimestamp)
        modifyTimes = get_timestr(modifyTimestamp)
        accessTimes = get_timestr(accessTimestamp)
        # print(createTimes)  # 格式 2023-03-10 17:16:27
        SetFileTime(fh, createTimes, accessTimes, modifyTimes)
        CloseHandle(fh)
        return True
    except:
        # print(e)
        CloseHandle(fh)#防止句柄泄漏造成操作系统出现隐性bug(例如文件部分属性无法成功设置之类的
        return False


def timeOffsetAndStruct(times, format, offset):
    """时间偏移"""
    return time.localtime(time.mktime(time.strptime(times, format)) + offset)


def getFileVersion(file_name):
    """获取文件版本号"""
    info = win32api.GetFileVersionInfo(file_name, os.sep)
    ms = info['FileVersionMS']
    ls = info['FileVersionLS']
    version = '%d.%d.%d.%d' % (win32api.HIWORD(ms), win32api.LOWORD(ms), win32api.HIWORD(ls), win32api.LOWORD(ls))
    return version


def export_media_info_as_excel_by_openpyxl(file_name, result):
    """使用openpyxl写出excel文件
    :param file_name: excel路径
    :param result: 要写出的数据
    """
    # 打开文件
    wb = openpyxl.Workbook()

    # 初始化字体对象
    font_head = openpyxl.styles.Font(
        size=12,
        bold=True,
    )
    # 初始化单元格对齐方式的对象
    alight_center = openpyxl.styles.Alignment(
        horizontal='center',  # 水平对齐方式:center, left, right
        vertical='center'  # 垂直对齐方式: center, top, bottom
    )
    alight_left = openpyxl.styles.Alignment(
        horizontal='left',  # 水平对齐方式:center, left, right
        vertical='center'  # 垂直对齐方式: center, top, bottom
    )
    alight_right = openpyxl.styles.Alignment(
        horizontal='right',  # 水平对齐方式:center, left, right
        vertical='center'  # 垂直对齐方式: center, top, bottom
    )

    # 初始化单元格样式信息
    default_width = 20
    default_height = 18
    default_alignment = alight_left
    # 初始化单元格样式配置
    STYLE_DICT = {
        "文件名": {"width": 30},
        "文件路径": {"width": 60},
        "文件大小": {"alignment": alight_right},
        "拍摄时间": {"width": 22},
        "GPS定位": {"width": 30},
        "创建日期": {"width": 30},
        "最近修改日期": {"width": 30},
        "帧速率": {"alignment": alight_right},
        "总比特率": {"alignment": alight_right},
        "音频采样率": {"alignment": alight_right},
        "音频流大小": {"alignment": alight_right},
        "音频比特率": {"alignment": alight_right},
    }

    # 写出数据库所有数据到excel
    for sheet_name, infos_list in result.items():
        sh = wb.create_sheet(sheet_name)

        # 获取表单的表头信息
        if len(infos_list) == 0:  # 表无数据则不创建
            continue
        titles = list(infos_list[0].keys())
        
        # 插入表头
        col = 1
        row = 1
        for title in titles:
            # 计算合适列宽
            cell = sh.cell(row=row, column=col, value=title)
            cell.font = font_head  # 设置首行字体
            cell.alignment = alight_center  # 设置首行对齐
            col += 1

        # 设置首行单元格高度
        sh.row_dimensions[1].height = 20

        # 设置首行单元格宽度
        for i in range(1, sh.max_column + 1):
            style_dict = STYLE_DICT.get(titles[i-1])
            width = default_width
            if style_dict:
                width = style_dict.get('width')
            width = width if width else default_width
            # print("%s 设置宽度: %s" % (titles[i-1], width))
            sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # 插入数据
        for info in infos_list:
            row += 1
            col = 1
            for title in titles:
                cell = sh.cell(row=row, column=col, value=info[title])
                sh.row_dimensions[row].height = default_height  # 设置单元格高度
                # 设置单元格对齐方式
                alignment = default_alignment
                if STYLE_DICT.get(title):
                    alignment = STYLE_DICT.get(title).get("alignment")
                cell.alignment = alignment if alignment else default_alignment
                col += 1

    wb.remove(wb["Sheet"])  # 删除默认创建的空工作簿
    # 保存到文件并关闭工作簿
    wb.save(file_name)
    wb.close()


